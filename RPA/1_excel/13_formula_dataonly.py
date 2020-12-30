from openpyxl import load_workbook
# wb = load_workbook("sample_fomula.xlsx")
# ws = wb.active
#
# #   수식 그대로 가져오고 있음
# for row in ws.values:
#     for cell in row:
#         print(cell)

wb = load_workbook("sample_fomula.xlsx", data_only=True)
ws = wb.active

# 수식이 아닌 실제 데이터를 가져오고 있음
# 수식은 파일 저장하기 전에는 값이 안나오고 None으로 뜸
for row in ws.values:
    for cell in row:
        print(cell)

# for row in ws.values:
#     print(row[0])

wb.save("sample_fomula.xlsx")