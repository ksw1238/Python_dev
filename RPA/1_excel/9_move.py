from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호 영어 수학
# 번호 (국어) 영어 수학

# ws.move_range("B1:C11", rows=0, cols=1)
# ws["B1"].value = "국어"   # B1 셀에 '국어' 입력

# 번호 수학 영어
ws.move_range("B1:B11", rows=0, cols=2)
ws.delete_cols(2)

wb.save("sample_move.xlsx")