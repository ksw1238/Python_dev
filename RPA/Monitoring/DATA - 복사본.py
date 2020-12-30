import os
import shutil
import time
import datetime
import pyautogui
from openpyxl import load_workbook   # 파일 불러오기
import win32com.client
import tkinter.messagebox as msgbox

# 중복제거 & copy
def duplicated(src_ws, src_col, tar_ws, tar_col):   # 소스 worksheet, 소스 열, 복사대상 worksheet, 복사대상 열
    try:
        my_list = []
        for row in src_ws.values:
            my_list.append(row[src_col-1])
        my_list.pop(0)   # 1행(제목) 삭제
        my_set = set(my_list)
        new_list = list(my_set)

        for x in range(1, len(new_list) + 1):
            tar_ws.cell(row=x, column=tar_col).value = new_list[x - 1]

    except Exception as err:   # 예외처리
        msgbox.showerror("에러", err)

# 모니터링 월의 마지막 날짜 가져오기
def getLastDay(year, month):   # 모니터링 대상 연도, 모니터링 대상 월
    try:
        if month == 12:
            year = year + 1
            month = 1
        else:
            month = month + 1
        d = datetime.date(year, month, 1)
        t = datetime.timedelta(days=1)
        k = d - t
        return k.day

    except Exception as err:   # 예외처리
        msgbox.showerror("에러", err)

# 파일 복사
def file_copy(src_path, tar_path, src, tar):   # 소스경로, 타겟경로, 소스파일명, 타켓파일명
    try:
        # target 경로가 존재하는지?
        if not os.path.exists(tar_path):
            print("존재하지 않습니다. 이전 폴더를 카피합니다.")
            target_path2 = "\\\lginas.lgcloud.com\lgi_cd$\경영정보팀\작업폴더\\2. 공통 작업 자료\C05_IT진단_5년\C_" + str(mon_year) + \
                           "년_5년\활용이력 모니터링\\5. 데이터 변경 모니터링 - 월별\\" + str(mon_year) + "." + str(mon_month - 1) + "\\"
            shutil.copytree(target_path2, tar_path)  # 폴더가 없을 경우 전월 폴더(target_path2)를 카피해서 신규 생성(taget_path)

        src_file = src_path + src
        tar_file = tar_path + tar

        # source 파일이 존재하는지? (존재하지 않으면 카피 skip)
        if not os.path.exists(src_file):
            msgbox.showinfo("알림", "'" + src + "'" + " 파일이 존재하지 않습니다.")
            print("'" + src + "'" + " 파일이 존재하지 않습니다.")
            return

        shutil.copy(src_file, tar_file)  # 파일명을 바꾸면서 copy(존재하는경우 덮어쓰기!!!)

    except Exception as err:   # 예외처리
        msgbox.showerror("에러", err)

# 모니터링 대상은 오늘 기준으로 1달 전
today = datetime.date.today()
mon_month = today.month - 1
if mon_month == 12:
    mon_year = today.year - 1
else:
    mon_year = today.year

# 모니터링 취합 경로 및 파일명 지정
source_path = "\\\lginas.lgcloud.com\\lgi_cd$\\경영정보팀\작업폴더\\2. 공통 작업 자료\\▲정보보안\\02_정보보안 증적자료\\9. 정보 기술" \
            "\\9.6 로그 관리\\로그관리현황\\" + str(mon_year) + "\\" + str(mon_month) + "월\\DB접근제어\\수신내역\\"
target_path = "\\\lginas.lgcloud.com\\lgi_cd$\\경영정보팀\\작업폴더\\2. 공통 작업 자료\\C05_IT진단_5년\\C_" + str(mon_year) + \
              "년_5년\\활용이력 모니터링\\5. 데이터 변경 모니터링 - 월별\\" + str(mon_year) + "." + str(mon_month) + "\\"

### JOYBILL 모니터링 ###
# 모니터링 대상 파일 가져오기
# (매달 지웅준 책임님이 담당자들에게 이메일로 요청해서 CSR번호 매칭시킨 파일을 저장함)
def Monitoring_JOYBILL():
    try:
        file_name = "황천규.xlsx"
        target_file = "'" + str(mon_year)[2:4] + "." + str(mon_month) + " Application 데이터 변경 점검_JOYBILL.xlsx"
        file_copy(source_path, target_path, file_name, target_file)   # 파일카피 - 황천규 선임 파일
        wb = load_workbook(target_path + target_file)
        ws = wb.active
        ws2 = wb.create_sheet('CSR_list')

        duplicated(ws, 48, ws2, 1)   # CSR번호 중복제거

        wb.save(target_path + target_file)   # 파일 저장

        msgbox.showinfo("알림", "JOYBILL 작업이 완료되었습니다.")

    except Exception as err:   # 예외처리
        msgbox.showerror("에러", err)

### e-HR 모니터링 ###
# 모니터링 대상 파일 가져오기
# (매달 지웅준 책임님이 담당자들에게 이메일로 요청해서 CSR번호 매칭시킨 파일을 저장함)
def Monitoring_HR():
    try:
        file_name = "이주연.xlsx"
        target_file = file_name
        file_copy(source_path, target_path, file_name, target_file)   # 파일카피 - 이주연 과장 파일
        wb_lee = load_workbook(target_path + target_file)
        ws_lee = wb_lee.active
        ws2_lee = wb_lee.create_sheet('CSR_list')

        file_name = "최연정.xlsx"
        target_file = file_name
        file_copy(source_path, target_path, file_name, target_file)   # 파일카피 - 최연정 사원 파일
        wb_choi = load_workbook(target_path + target_file)
        ws_choi = wb_choi.active

        # 2개 파일 합체
        for row in ws_choi.iter_rows(min_row=2, max_row=ws_choi.max_row, min_col=1, max_col=ws_choi.max_column):
            a = []
            for cell in row:
                a.append(cell.value)
            ws_lee.append(a)

        duplicated(ws_lee, 48, ws2_lee, 1)   # CSR번호 중복제거

        target_file = "'" + str(mon_year)[2:4] + "." + str(mon_month) + " Application 데이터 변경 점검_HR.xlsx"
        wb_lee.save(target_path + target_file)   # 파일 저장

        msgbox.showinfo("알림", "HR 작업이 완료되었습니다.")

    except Exception as err:   # 예외처리
        msgbox.showerror("에러", err)

### SAP 모니터링 ###
# 프로그램 실행
def Monitoring_SAP(ID, PW):
    try:
        pyautogui.press('win', interval=1)
        pyautogui.write("sap")
        pyautogui.press('enter', interval=8)
        w = pyautogui.getWindowsWithTitle('SAP Logon 760')[0]
        w.activate()
        pyautogui.write(['down', 'down', 'down'])
        pyautogui.press('enter', interval=3)


        # 로그인
        pyautogui.write(ID)
        pyautogui.press('tab')
        pyautogui.write(PW)
        pyautogui.press('enter', interval=1)
        pyautogui.write("2190033")
        pyautogui.press('tab')
        pyautogui.write(PW)
        pyautogui.press('enter', interval=2)
        pyautogui.write("se16")
        pyautogui.press('enter', interval=1)

        # SE16N_CD_KEY 테이블 데이터 다운로드
        pyautogui.write("SE16N_CD_KEY")
        pyautogui.press('enter', interval=1)
        pyautogui.write(['tab', 'tab', 'tab', 'tab', 'tab', 'tab', 'tab', 'tab', 'tab'])
        valid_from = str(mon_year) + str(mon_month) + '01'
        pyautogui.write(valid_from)
        pyautogui.press('tab')
        last_day = getLastDay(mon_year, mon_month)   # 모니터링 월의 마지막 날짜 가져오기
        valid_to = str(mon_year) + str(mon_month) + str(last_day)
        pyautogui.write(valid_to)
        pyautogui.write(['tab', 'tab', 'tab'])
        pyautogui.write("24")
        pyautogui.press('enter')
        pyautogui.write(['tab', 'tab', 'tab', 'tab', 'tab', 'tab'])
        pyautogui.press('del')
        pyautogui.press('f8', interval=1)
        pyautogui.hotkey("ctrl", "shift", "f7", interval=0.25)
        pyautogui.press('enter', interval=2)
        pyautogui.hotkey("alt", "s")
        time.sleep(1)
        pyautogui.hotkey("alt", "y")
        time.sleep(2)
        pyautogui.hotkey("alt", "a")
        time.sleep(4)
        pyautogui.hotkey("alt", "f4", interval=1)
        SAP_source_path = "C:/Users/khkimb/Documents/SAP/SAP GUI/"
        file_name = "export.MHTML"
        target_file = "SE16N_CD_KEY.MHTML"
        file_copy(SAP_source_path, target_path, file_name, target_file)   # 파일카피 - SE16N_CD_KEY 파일

        # SE16N_CD_DATA 테이블 데이터 다운로드
        pyautogui.press('right')
        pyautogui.hotkey("ctrl", "space")
        pyautogui.hotkey("ctrl", "c")
        pyautogui.write(['f3', 'f3'], interval=0.25)
        pyautogui.write("SE16N_CD_DATA")
        pyautogui.press('enter', interval=1)
        pyautogui.write(['tab', 'tab', 'space'], interval=0.25)
        pyautogui.hotkey("SHIFT", "f12", interval=0.25)
        pyautogui.press('f8', interval=0.25)
        pyautogui.write(['tab', 'tab', 'tab', 'tab', 'tab', 'tab', 'tab', 'tab', 'tab', 'tab', 'tab', 'tab', 'tab', 'tab', 'tab'])
        pyautogui.press('del')
        pyautogui.press('f8', interval=1)
        pyautogui.hotkey("ctrl", "shift", "f7", interval=0.25)
        pyautogui.press('enter', interval=3)
        pyautogui.hotkey("alt", "s")
        time.sleep(1)
        pyautogui.hotkey("alt", "y")
        time.sleep(4)
        pyautogui.hotkey("alt", "a")
        time.sleep(6)
        pyautogui.hotkey("alt", "f4", interval=1)
        target_file = "SE16N_CD_DATA.MHTML"
        file_copy(SAP_source_path, target_path, file_name, target_file)   # 파일카피 - SE16N_CD_KEY 파일

        # 데이터 수정 권한 관리 데이터 복사 - ZAUR0400
        pyautogui.write(['f3', 'f3', 'tab', 'tab'], interval=0.25)
        pyautogui.write("/nZAUR0400")
        pyautogui.press('enter', interval=1)
        pyautogui.write(valid_from)
        pyautogui.press('tab')
        pyautogui.write(valid_to)
        pyautogui.write(['tab', 'tab', 'tab', 'tab', 'tab', 'tab', 'tab', 'tab', 'up', 'up', 'f8'], interval=0.25)
        time.sleep(1)
        pyautogui.rightClick(500, 400, duration=1)
        time.sleep(2)
        pyautogui.press('up', interval=0.25)
        pyautogui.press('enter', interval=0.25)
        pyautogui.press('enter', interval=0.25)
        time.sleep(3)
        pyautogui.hotkey("alt", "s")
        time.sleep(1)
        pyautogui.hotkey("alt", "y")
        time.sleep(2)
        pyautogui.hotkey("alt", "a")
        time.sleep(4)
        pyautogui.hotkey("alt", "f4", interval=1)
        target_file = "ZAUR0400.MHTML"
        file_copy(SAP_source_path, target_path, file_name, target_file)   # 파일카피 - ZAUR0400 파일

        # SE16N_CD_KEY & SE16N_CD_DATA 데이터 합치기
        excelApp = win32com.client.Dispatch('Excel.Application')
        excelApp.DisplayAlerts = False   # 덮어쓰기 경고창 끄기(자동 덮어쓰기)
        book = excelApp.Workbooks.Open(target_path + "SE16N_CD_KEY.MHTML")
        book.SaveAs(target_path + "SE16N_CD_KEY.xlsx", 51)
        book.Close()
        book = excelApp.Workbooks.Open(target_path + "SE16N_CD_DATA.MHTML")
        book.SaveAs(target_path + "SE16N_CD_DATA.xlsx", 51)
        book.Close()
        book = excelApp.Workbooks.Open(target_path + "ZAUR0400.MHTML")
        book.SaveAs(target_path + "ZAUR0400.xlsx", 51)
        book.Close()

        source_file = "'" + str(mon_year)[2:4] + "." + str(mon_month-1) + " Application 데이터 변경 점검_SAP.xlsx"
        target_wb = load_workbook(target_path + source_file, data_only=True)   # 전월 데이터 카피 떠온걸 연다
        target_ws = target_wb.active
        target_ws2 = target_wb.worksheets[1]
        target_ws3 = target_wb.create_sheet('ZAUR0400')
        KEY_wb = load_workbook(target_path + "SE16N_CD_KEY.xlsx")
        KEY_ws = KEY_wb.active
        DATA_wb = load_workbook(target_path + "SE16N_CD_DATA.xlsx")
        DATA_ws = DATA_wb.active
        SAP_wb = load_workbook(target_path + "ZAUR0400.xlsx")
        SAP_ws = SAP_wb.active

        target_ws.delete_rows(6, target_ws.max_row+1)   # sheet1 데이터만 삭제
        target_ws2.delete_rows(1, target_ws2.max_row+1)   # sheet2 전체 삭제
        target_ws3.delete_rows(1, target_ws3.max_row+1)   # sheet3 전체 삭제


        for x in range(2, DATA_ws.max_row+1):
            for y in range(1, 7):
                target_ws.cell(row=x+4, column=y).value = DATA_ws.cell(row=x, column=y).value   # 데이터 복사

        for x in range(2, DATA_ws.max_row+1):
                target_ws.cell(row=x+4, column=11).value = DATA_ws.cell(row=x, column=7).value   # Data part 복사

        # SE16N_CD_KEY.xlsx 파일 데이터를 sheet2에 복사
        for x in range(1, KEY_ws.max_row+1):
            for y in range(1, 8):
                target_ws2.cell(row=x, column=y).value = KEY_ws.cell(row=x, column=y).value

        for x in range(2, DATA_ws.max_row+1):
            target_ws['G' + str(x+4)] = "=VLOOKUP(B" + str(x+4) + ",Sheet2!B:E,4,0)"   # SE16N_CD_KEY.xlsx에서 날짜 복사
            target_ws.cell(row=x+4, column=7).number_format = 'YYYY/MM/DD'

        for x in range(2, DATA_ws.max_row+1):
            target_ws['H' + str(x+4)] = "=VLOOKUP(B" + str(x+4) + ",Sheet2!B:E,3,0)"   # SE16N_CD_KEY.xlsx에서 User ID 복사

        # ZAUR0400.xlsx 파일 데이터를 sheet3에 복사
        for x in range(1, SAP_ws.max_row+1):
            for y in range(1, SAP_ws.max_column+1):
                target_ws3.cell(row=x, column=y).value = SAP_ws.cell(row=x, column=y).value

        # CSR번호 자동으로 찾아오기(못찾은 것들 수기로 확인 필요!!)
        for x in range(2, DATA_ws.max_row+1):
            target_ws['C' + str(x+4)] = "=CONCATENATE(H" + str(x+4) + ",G" + str(x+4) + ")"   # Contcatinate(UserID, 날짜)

        for x in range(2, SAP_ws.max_row+1):
            target_ws3['A' + str(x)] = "=CONCATENATE(D" + str(x) + ",L" + str(x) + ")"   # Contcatinate(UserID, 날짜)

        for x in range(2, DATA_ws.max_row+1):
            target_ws['I' + str(x+4)] = "=VLOOKUP(C" + str(x+4) + ",ZAUR0400!A:H,8,0)"   # CSR번호 붙이기

        target_file = "'" + str(mon_year)[2:4] + "." + str(mon_month) + " Application 데이터 변경 점검_SAP.xlsx"
        target_wb.save(target_path + target_file)

        # 수식을 값으로 처리하기 위해 win32com으로 저장
        book = excelApp.Workbooks.Open(target_path + target_file)
        book.SaveAs(target_path + target_file, 51)
        book.Close()

        wb = load_workbook(target_path + target_file, data_only=True)   # 전월 데이터 카피 떠온걸 연다
        ws = wb.worksheets[0]
        ws4 = wb.create_sheet('CSR_list')

        duplicated(ws, 9, ws4, 1)   # CSR번호 중복제거

        wb.save(target_path + target_file)

        msgbox.showinfo("알림", "SAP 작업이 완료되었습니다.")

    except Exception as err:   # 예외처리
        msgbox.showerror("에러", err)